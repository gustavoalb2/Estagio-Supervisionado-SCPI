from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from .models import Processo, Usuario, TabelaProcessos, Auditoria
from .forms import ProcessoForm, TabelaForm, AlterarSenhaPropegForm
from datetime import datetime
from django.db.models import Q
import json

# csv, xlsx
import csv
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# Funções auxiliares para auditoria
def registrar_auditoria(usuario, acao, processo=None, tabela=None, detalhes=None):
    try:
        Auditoria.objects.create(
            usuario=usuario,
            acao=acao,
            processo=processo,
            tabela=tabela,
            detalhes=detalhes
        )
    except Exception as e:
        print(f"Erro ao registrar auditoria: {e}")

def gerar_detalhes_processo(processo, acao):
    detalhes = {
        'acao': acao,
        'processo_info': {
            'id': processo.id,
            'nome': processo.nome,
            'matricula': processo.matricula,
            'numero_processo': processo.numero_processo,
            'setor': processo.setor,
            'bolsa': processo.bolsa,
            'status': processo.status,
            'tabela_id': processo.tabela.id if processo.tabela else None,
            'tabela_nome': processo.tabela.nome if processo.tabela else None
        }
    }
    return json.dumps(detalhes, ensure_ascii=False)

def gerar_detalhes_tabela(tabela, acao):
    detalhes = {
        'acao': acao,
        'tabela_info': {
            'id': tabela.id,
            'nome': tabela.nome,
            'descricao': tabela.descricao,
            'data_criacao': str(tabela.data_criacao)
        }
    }
    return json.dumps(detalhes, ensure_ascii=False)

@login_required(login_url='login')
def home(request):
    print(f"DEBUG - Usuário autenticado: {request.user.is_authenticated}")
    print(f"DEBUG - Usuário: {request.user}")
    
    query = request.GET.get('q')
    if query:
        tabelas = TabelaProcessos.objects.filter(nome__icontains=query)
    else:
        tabelas = TabelaProcessos.objects.all()
    
    context = {
        'tabelas': tabelas
    }
    return render(request, 'visualizarTabelas.html', context)

def processo(request):
    processos = TabelaProcessos.objects.all()
    return render(request, 'tabelaProcessos.html', {'processos': processos})

def adicionarTabela(request):
    if request.method == 'POST':
        try:
            nome_tabela = request.POST.get('titulo')
            descricao_tabela = request.POST.get('descricao')
            
            usuario, created = Usuario.objects.get_or_create(
                id=1, 
                defaults={'nome': 'Usuário Padrão', 'email': 'padrao@example.com'}
            )

            if not nome_tabela:
                messages.error(request, 'O título da tabela é obrigatório.')
                return render(request, 'adicionarTabela.html')

            nova_tabela = TabelaProcessos(
                nome=nome_tabela,
                descricao=descricao_tabela,
                usuario=usuario
            )
            nova_tabela.save()
            
            # Registrar auditoria
            detalhes = gerar_detalhes_tabela(nova_tabela, 'CRIAR')
            registrar_auditoria(
                usuario=request.user,
                acao='CRIAR',
                tabela=nova_tabela,
                detalhes=detalhes
            )
            
            messages.success(request, f'Tabela "{nome_tabela}" criada com sucesso!')
            return redirect('visualizarTabelas') # Redireciona para a página inicial ou de visualização de tabelas

        except Exception as e:
            messages.error(request, f'Erro ao criar a tabela: {str(e)}')
            
    return render(request, 'adicionarTabela.html')

def editarProcesso(request, processo_id):
    processo = get_object_or_404(Processo, id=processo_id)
    if request.method == 'POST':
        form = ProcessoForm(request.POST, instance=processo)
        if form.is_valid():
            processo_atualizado = form.save()
            
            # Registrar auditoria
            detalhes = gerar_detalhes_processo(processo_atualizado, 'ATUALIZAR')
            registrar_auditoria(
                usuario=request.user,
                acao='ATUALIZAR',
                processo=processo_atualizado,
                tabela=processo_atualizado.tabela,
                detalhes=detalhes
            )
            
            messages.success(request, f'Processo "{processo.numero_processo}" atualizado com sucesso!')
            if processo.tabela:
                return redirect('tabela_processos', tabela_id=processo.tabela.id)
            else:
                return redirect('visualizarTabelas')
        else:
            messages.error(request, 'Erro ao atualizar o processo. Verifique os dados informados.')
    else:
        form = ProcessoForm(instance=processo)

    context = {
        'form': form,
        'tabela': processo.tabela,
        'processo': processo  # Passa o objeto processo para o template
    }
    return render(request, 'processo_form.html', context)

def deletaProcesso(request, processo_id):
    processo = get_object_or_404(Processo, id=processo_id)
    tabela_id = processo.tabela.id if processo.tabela else None

    if request.method == 'POST':
        try:
            numero_processo = processo.numero_processo
            detalhes = gerar_detalhes_processo(processo, 'EXCLUIR')
            registrar_auditoria(
                usuario=request.user,
                acao='EXCLUIR',
                processo=None,  # Será None pois o processo será deletado
                tabela=processo.tabela,
                detalhes=detalhes
            )
            
            processo.delete()
            messages.success(request, f'Processo {numero_processo} deletado com sucesso!')
        except Exception as e:
            messages.error(request, f'Erro ao deletar processo: {str(e)}')
    
    if tabela_id:
        return redirect('tabela_processos', tabela_id=tabela_id)
    else:
        return redirect('visualizarTabelas')

def adicionarProcesso(request, tabela_id=None):
    tabela = None
    if tabela_id:
        tabela = get_object_or_404(TabelaProcessos, pk=tabela_id)

    if request.method == 'POST':
        form = ProcessoForm(request.POST)
        if form.is_valid():
            processo = form.save(commit=False)
            if tabela:
                processo.tabela = tabela
            processo.save()
            
            # Registrar auditoria
            detalhes = gerar_detalhes_processo(processo, 'CRIAR')
            registrar_auditoria(
                usuario=request.user,
                acao='CRIAR',
                processo=processo,
                tabela=processo.tabela,
                detalhes=detalhes
            )
            
            messages.success(request, f'Processo "{processo.numero_processo}" adicionado com sucesso!')
            if tabela:
                return redirect('tabela_processos', tabela_id=tabela.id)
            else:
                return redirect('visualizarTabelas')
        else:
            messages.error(request, 'Erro ao adicionar o processo. Verifique os dados informados.')
    else:
        from datetime import date
        initial_data = {'data_abertura': date.today()}
        form = ProcessoForm(initial=initial_data)

    context = {
        'form': form,
        'tabela': tabela
    }
    return render(request, 'processo_form.html', context)

def tabela(request, tabela_id):
    tabela = get_object_or_404(TabelaProcessos, id=tabela_id)
    query = request.GET.get('q')
    
    # Parâmetros de ordenação
    sort_by = request.GET.get('sort', 'nome')  # Ordenação padrão por nome
    sort_direction = request.GET.get('direction', 'asc')  # Direção padrão ascendente
    
    if query:
        processos = Processo.objects.filter(
            Q(tabela=tabela) & 
            (Q(nome__icontains=query) | 
             Q(numero_processo__icontains=query) |
             Q(assunto__icontains=query))
        )
    else:
        processos = Processo.objects.filter(tabela=tabela)
    
    # Aplicar ordenação
    if sort_by == 'nome':
        if sort_direction == 'desc':
            processos = processos.order_by('-nome')
        else:
            processos = processos.order_by('nome')
    elif sort_by == 'data_abertura':
        if sort_direction == 'desc':
            processos = processos.order_by('-data_abertura')
        else:
            processos = processos.order_by('data_abertura')
    elif sort_by == 'data_retorno':
        if sort_direction == 'desc':
            processos = processos.order_by('-data_retorno')
        else:
            processos = processos.order_by('data_retorno')
    
    # Contagem de processos por setor
    count_cic = processos.filter(setor='CIC').count()
    count_dpq = processos.filter(setor='DPQ').count()
    
    # Contagem total para exibir informação sobre filtragem
    total_processos = processos.count()
    total_sem_filtro = Processo.objects.filter(tabela=tabela).count()
    tem_filtro = query is not None and query != ''

    context = {
        'processos': processos,
        'tabela': tabela,
        'count_cic': count_cic,
        'count_dpq': count_dpq,
        'sort_by': sort_by,
        'sort_direction': sort_direction,
        'total_processos': total_processos,
        'total_sem_filtro': total_sem_filtro,
        'tem_filtro': tem_filtro,
    }
    return render(request, 'tabelaProcessos.html', context)

@login_required(login_url='login')
def usuarios(request):
    # Pega o usuário logado atualmente
    current_user = request.user
    
    nome_completo = current_user.get_full_name()
    if not nome_completo or nome_completo.strip() == '':
        nome_completo = current_user.first_name or current_user.username.upper()
        
    email_usuario = current_user.email or f'{current_user.username}@ufac.br'
    
    usuario_info = {
        'nome': nome_completo,
        'email': email_usuario,
        'username': current_user.username,
        'data_criacao': current_user.date_joined,
        'ultimo_login': current_user.last_login,
        'is_active': current_user.is_active
    }
    
    context = {
        'usuario_info': usuario_info,
        'is_admin': current_user.is_superuser  
    }
    return render(request, 'usuario.html', context)

def editar_tabela(request, tabela_id):
    tabela = get_object_or_404(TabelaProcessos, id=tabela_id)
    if request.method == 'POST':
        form = TabelaForm(request.POST, instance=tabela)
        if form.is_valid():
            tabela_atualizada = form.save()
            
            # Registrar auditoria
            detalhes = gerar_detalhes_tabela(tabela_atualizada, 'ATUALIZAR')
            registrar_auditoria(
                usuario=request.user,
                acao='ATUALIZAR',
                tabela=tabela_atualizada,
                detalhes=detalhes
            )
            
            messages.success(request, f'Tabela "{tabela.nome}" atualizada com sucesso!')
            return redirect('visualizarTabelas')
        else:
            messages.error(request, 'Erro ao atualizar a tabela. Verifique os dados informados.')
    else:
        form = TabelaForm(instance=tabela)

    context = {
        'form': form,
        'tabela': tabela
    }
    return render(request, 'editar_tabela.html', context)

def excluir_tabela(request, tabela_id):
    tabela = get_object_or_404(TabelaProcessos, id=tabela_id)
    if request.method == 'POST':
        try:
            nome_tabela = tabela.nome
            
            # Registrar auditoria antes de deletar
            detalhes = gerar_detalhes_tabela(tabela, 'EXCLUIR')
            registrar_auditoria(
                usuario=request.user,
                acao='EXCLUIR',
                tabela=None,  # Será None pois a tabela será deletada
                detalhes=detalhes
            )
            
            tabela.delete()
            messages.success(request, f'Tabela "{nome_tabela}" excluída com sucesso!')
        except Exception as e:
            messages.error(request, f'Erro ao excluir a tabela: {str(e)}')
        return redirect('visualizarTabelas')
    
    context = {
        'tabela': tabela
    }
    return render(request, 'confirmar_exclusao_tabela.html', context)

def exportar_xlsx(request, tabela_id):
    tabela = get_object_or_404(TabelaProcessos, id=tabela_id)
    query = request.GET.get('q')
    
    # Parâmetros de ordenação
    sort_by = request.GET.get('sort', 'nome')  # Ordenação padrão por nome
    sort_direction = request.GET.get('direction', 'asc')  # Direção padrão ascendente
    
    # Filtrar processos de acordo com a busca, se houver
    if query:
        processos = Processo.objects.filter(
            Q(tabela=tabela) & 
            (Q(nome__icontains=query) | 
             Q(numero_processo__icontains=query) |
             Q(assunto__icontains=query))
        )
    else:
        processos = Processo.objects.filter(tabela=tabela)
    
    # Aplicar ordenação
    if sort_by == 'nome':
        if sort_direction == 'desc':
            processos = processos.order_by('-nome')
        else:
            processos = processos.order_by('nome')
    elif sort_by == 'data_abertura':
        if sort_direction == 'desc':
            processos = processos.order_by('-data_abertura')
        else:
            processos = processos.order_by('data_abertura')
    elif sort_by == 'data_retorno':
        if sort_direction == 'desc':
            processos = processos.order_by('-data_retorno')
        else:
            processos = processos.order_by('data_retorno')
    
    # Criar um novo workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Processos"
    
    # Definir estilos
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='006600', end_color='006600', fill_type='solid')
    
    # Definir bordas
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Alinhamento
    center_aligned = Alignment(horizontal='center', vertical='center')
    wrapped_alignment = Alignment(wrap_text=True, vertical='top')
    
    # Definir estilo zebra para linhas alternadas
    light_green_fill = PatternFill(start_color='E6FFE6', end_color='E6FFE6', fill_type='solid')
    
    # Adicionar título principal
    ws.merge_cells('A1:J1')  # Mesclar células para o título
    titulo_cell = ws['A1']
    
    # Adicionar informações sobre filtro e ordenação no título
    sort_by_info = ""
    if sort_by:
        # Mapeamento de campos para exibição amigável
        campo_exibicao = {
            'nome': 'Nome',
            'data_abertura': 'Data de Abertura',
            'data_retorno': 'Data de Retorno'
        }
        
        # Mapeamento de direção para exibição amigável
        direcao_exibicao = {
            'asc': 'crescente',
            'desc': 'decrescente'
        }
        
        sort_by_info = f" - Ordenado por {campo_exibicao.get(sort_by, sort_by)} ({direcao_exibicao.get(sort_direction, sort_direction)})"
    
    if query:
        titulo_cell.value = f"SCPI - Processos filtrados da Tabela: {tabela.nome} (Filtro: {query}){sort_by_info}"
    else:
        titulo_cell.value = f"SCPI - Processos da Tabela: {tabela.nome}{sort_by_info}"
    
    titulo_cell.font = Font(name='Arial', size=14, bold=True)
    titulo_cell.alignment = center_aligned
    titulo_cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
    
    # Adicionar cabeçalhos na linha 2
    headers = ['Nome', 'Matrícula', 'Nº Processo', 'Data de Abertura', 'Data de Retorno', 'Setor', 'Bolsa', 'Status', 'Assunto', 'Observações']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_aligned
    
    # Adicionar dados
    for row_num, processo in enumerate(processos, 3):  # Começar da linha 3 (após título e cabeçalho)
        # Aplicar cor de fundo alternada
        row_fill = light_green_fill if row_num % 2 == 0 else None
        
        # Formatar data para exibição
        data_abertura = processo.data_abertura.strftime('%d/%m/%Y') if processo.data_abertura else ""
        data_retorno = processo.data_retorno.strftime('%d/%m/%Y') if processo.data_retorno else ""
        
        # Preencher células com valores e aplicar formatação
        cells = [
            (1, processo.nome),
            (2, processo.matricula),
            (3, processo.numero_processo),
            (4, data_abertura),
            (5, data_retorno),
            (6, processo.get_setor_display() if processo.setor else ""),
            (7, processo.get_bolsa_display() if processo.bolsa else ""),
            (8, processo.get_status_display() if processo.status else ""),
            (9, processo.assunto),
            (10, processo.observacoes)
        ]
        
        for col_num, value in cells:
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            
            # Aplicar alinhamento e quebra de texto para campos longos
            if col_num in [9, 10]:  # Assunto e Observações
                cell.alignment = wrapped_alignment
            else:
                cell.alignment = Alignment(vertical='center')
                
            # Aplicar cor de fundo alternada
            if row_fill:
                cell.fill = row_fill
    
    # Ajustar largura das colunas
    column_widths = {
        'A': 25,  # Nome
        'B': 15,  # Matrícula
        'C': 20,  # Nº Processo
        'D': 15,  # Data de Abertura
        'E': 15,  # Data de Retorno
        'F': 15,  # Setor
        'G': 10,  # Bolsa
        'H': 15,  # Status
        'I': 30,  # Assunto
        'J': 30,  # Observações
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Ajustar altura das linhas
    for row_num in range(3, len(processos) + 3):
        ws.row_dimensions[row_num].height = 30  # Altura fixa para todas as linhas de dados
    
    # Congelar painel para manter cabeçalhos visíveis ao rolar
    ws.freeze_panes = 'A3'
    
    # Criar a resposta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="processos_{tabela.nome}.xlsx"'
    
    # Salvar o workbook para a resposta
    wb.save(response)
    
    return response

def exportar_processos_xlsx(request, tabela_id):
    response = exportar_xlsx(request, tabela_id)
    
    # Adicionar mensagem de confirmação com contagem
    tabela = get_object_or_404(TabelaProcessos, id=tabela_id)
    query = request.GET.get('q')
    sort_by = request.GET.get('sort', 'nome')
    sort_direction = request.GET.get('direction', 'asc')
    
    # Mapeamento de campos para exibição amigável
    campo_exibicao = {
        'nome': 'Nome',
        'data_abertura': 'Data de Abertura',
        'data_retorno': 'Data de Retorno'
    }
    
    # Mapeamento de direção para exibição amigável
    direcao_exibicao = {
        'asc': 'crescente',
        'desc': 'decrescente'
    }
    
    # Construir mensagem de sucesso
    ordenacao_info = f" (ordenados por {campo_exibicao.get(sort_by, sort_by)} em ordem {direcao_exibicao.get(sort_direction, sort_direction)})"
    
    if query:
        processos = Processo.objects.filter(
            Q(tabela=tabela) & 
            (Q(nome__icontains=query) | 
             Q(numero_processo__icontains=query) |
             Q(assunto__icontains=query))
        ).count()
        messages.success(request, f"{processos} processos filtrados foram exportados para Excel{ordenacao_info}.")
    else:
        processos = Processo.objects.filter(tabela=tabela).count()
        messages.success(request, f"Todos os {processos} processos foram exportados para Excel{ordenacao_info}.")
    
    return response

def exportar_processos_csv(request, tabela_id):
    tabela = get_object_or_404(TabelaProcessos, id=tabela_id)
    query = request.GET.get('q')
    
    # Parâmetros de ordenação
    sort_by = request.GET.get('sort', 'nome')  # Ordenação padrão por nome
    sort_direction = request.GET.get('direction', 'asc')  # Direção padrão ascendente
    
    # Filtrar processos de acordo com a busca, se houver
    # Mapeamento de campos para exibição amigável
    campo_exibicao = {
        'nome': 'Nome',
        'data_abertura': 'Data de Abertura',
        'data_retorno': 'Data de Retorno'
    }
    
    # Mapeamento de direção para exibição amigável
    direcao_exibicao = {
        'asc': 'crescente',
        'desc': 'decrescente'
    }
    
    # Construir mensagem de sucesso
    ordenacao_info = f" (ordenados por {campo_exibicao.get(sort_by, sort_by)} em ordem {direcao_exibicao.get(sort_direction, sort_direction)})"
    
    if query:
        processos = Processo.objects.filter(
            Q(tabela=tabela) & 
            (Q(nome__icontains=query) | 
             Q(numero_processo__icontains=query) |
             Q(assunto__icontains=query))
        )
        processos_count = processos.count()
        messages.success(request, f"{processos_count} processos filtrados foram exportados para CSV{ordenacao_info}.")
    else:
        processos = Processo.objects.filter(tabela=tabela)
        processos_count = processos.count()
        messages.success(request, f"Todos os {processos_count} processos foram exportados para CSV{ordenacao_info}.")
    
    # Aplicar ordenação
    if sort_by == 'nome':
        if sort_direction == 'desc':
            processos = processos.order_by('-nome')
        else:
            processos = processos.order_by('nome')
    elif sort_by == 'data_abertura':
        if sort_direction == 'desc':
            processos = processos.order_by('-data_abertura')
        else:
            processos = processos.order_by('data_abertura')
    elif sort_by == 'data_retorno':
        if sort_direction == 'desc':
            processos = processos.order_by('-data_retorno')
        else:
            processos = processos.order_by('data_retorno')
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="processos_{tabela.nome}.csv"'
    
    # Usar ponto-e-vírgula como delimitador para melhor compatibilidade com Excel brasileiro
    writer = csv.writer(response, delimiter=';', quoting=csv.QUOTE_ALL)
    
    # Adicionar BOM (Byte Order Mark) para garantir que o Excel interprete corretamente caracteres especiais
    response.write(u'\ufeff'.encode('utf-8'))
    
    # Adicionar linha com o nome da tabela
    # Mapeamento de campos para exibição amigável
    campo_exibicao = {
        'nome': 'Nome',
        'data_abertura': 'Data de Abertura',
        'data_retorno': 'Data de Retorno'
    }
    
    # Mapeamento de direção para exibição amigável
    direcao_exibicao = {
        'asc': 'crescente',
        'desc': 'decrescente'
    }
    
    sort_by_info = ""
    if sort_by:
        sort_by_info = f" - Ordenado por {campo_exibicao.get(sort_by, sort_by)} ({direcao_exibicao.get(sort_direction, sort_direction)})"
    
    if query:
        writer.writerow([f'SCPI - Processos filtrados da Tabela: {tabela.nome} (Filtro: {query}){sort_by_info}'])
    else:
        writer.writerow([f'SCPI - Processos da Tabela: {tabela.nome}{sort_by_info}'])
    writer.writerow([]) # Linha vazia
    
    # Adicionar cabeçalhos
    writer.writerow(['Nome', 'Matrícula', 'Nº Processo', 'Data de Abertura', 'Data de Retorno', 'Setor', 'Bolsa', 'Status', 'Assunto', 'Observações'])
    
    # Adicionar dados
    for processo in processos:
        # Formatar as datas para exibição no formato DD/MM/AAAA
        data_abertura = processo.data_abertura.strftime('%d/%m/%Y') if processo.data_abertura else ""
        data_retorno = processo.data_retorno.strftime('%d/%m/%Y') if processo.data_retorno else ""
        
        writer.writerow([
            processo.nome or "",
            processo.matricula or "",
            processo.numero_processo or "",
            data_abertura,
            data_retorno,
            processo.get_setor_display() if processo.setor else "",
            processo.get_bolsa_display() if processo.bolsa else "",
            processo.get_status_display() if processo.status else "",
            processo.assunto or "",
            processo.observacoes or ""
        ])
    
    return response

def importar_processos(request, tabela_id):
    if request.method == 'POST':
        tabela = get_object_or_404(TabelaProcessos, id=tabela_id)
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, "Nenhum arquivo foi enviado.")
            return redirect('tabela_processos', tabela_id=tabela.id)

        try:
            workbook = load_workbook(excel_file)
            sheet = workbook.active
            
            processos_importados = 0
            processos_ignorados = 0
            erros_detalhados = []
            
            # Determinar cabeçalhos e suas posições
            headers = None
            for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
                headers = row
                break
            
            # Definir índices padrão (baseados na ordem da planilha do contexto)
            nome_idx = 0
            matricula_idx = 1
            numero_processo_idx = 2
            data_abertura_idx = 3
            data_retorno_idx = 4
            setor_idx = 5
            bolsa_idx = 6
            status_idx = 7
            assunto_idx = 8
            observacoes_idx = 9
            
            # Se os cabeçalhos existirem, ajustar os índices
            if headers:
                # Mapear cabeçalhos para índices (caso a ordem da planilha seja diferente)
                header_map = {h.lower().strip() if h else "": i for i, h in enumerate(headers)}
                
                # Atualizar índices com base nos cabeçalhos encontrados
                nome_idx = header_map.get("nome", 0)
                matricula_idx = header_map.get("matrícula", 1)
                numero_processo_idx = header_map.get("nº processo", 2)
                data_abertura_idx = header_map.get("data de abertura", 3)
                data_retorno_idx = header_map.get("data de retorno", 4)
                setor_idx = header_map.get("setor", 5)
                bolsa_idx = header_map.get("bolsa", 6)
                status_idx = header_map.get("status", 7)
                assunto_idx = header_map.get("assunto", 8)
                observacoes_idx = header_map.get("observações", 9)
            
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                # Verificar se a linha tem dados suficientes
                if not row or len(row) < 1 or not row[nome_idx]:  # Verificar se pelo menos o nome existe
                    continue  # Pular linhas vazias ou sem nome
                
                # Extrair valores de forma segura com valores padrão para campos faltantes
                nome = row[nome_idx] if len(row) > nome_idx and row[nome_idx] else ""
                matricula = row[matricula_idx] if len(row) > matricula_idx and row[matricula_idx] else None
                numero_processo = row[numero_processo_idx] if len(row) > numero_processo_idx and row[numero_processo_idx] else None
                
                # Processamento de datas - convertendo para formato adequado ao Django
                data_abertura = None
                data_retorno = None
                
                try:
                    if len(row) > data_abertura_idx and row[data_abertura_idx]:
                        # Se for um objeto datetime do Excel
                        if hasattr(row[data_abertura_idx], 'date'):
                            data_abertura = row[data_abertura_idx].date()
                        # Se for uma string, tenta converter para data
                        elif isinstance(row[data_abertura_idx], str):
                            try:
                                # Tenta vários formatos comuns de data
                                for fmt in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y']:
                                    try:
                                        data_abertura = datetime.strptime(row[data_abertura_idx], fmt).date()
                                        break
                                    except ValueError:
                                        continue
                            except Exception:
                                pass
                except Exception as date_error:
                    erros_detalhados.append(f"Linha {row_num}: Erro ao processar data de abertura: {date_error}")

                try:
                    if len(row) > data_retorno_idx and row[data_retorno_idx]:
                        # Se for um objeto datetime do Excel
                        if hasattr(row[data_retorno_idx], 'date'):
                            data_retorno = row[data_retorno_idx].date()
                        # Se for uma string, tenta converter para data
                        elif isinstance(row[data_retorno_idx], str):
                            try:
                                # Tenta vários formatos comuns de data
                                for fmt in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y']:
                                    try:
                                        data_retorno = datetime.strptime(row[data_retorno_idx], fmt).date()
                                        break
                                    except ValueError:
                                        continue
                            except Exception:
                                pass
                except Exception as date_error:
                    erros_detalhados.append(f"Linha {row_num}: Erro ao processar data de retorno: {date_error}")
                
                setor = row[setor_idx] if len(row) > setor_idx and row[setor_idx] else None
                bolsa = row[bolsa_idx] if len(row) > bolsa_idx and row[bolsa_idx] else None
                status = row[status_idx] if len(row) > status_idx and row[status_idx] else None
                assunto = row[assunto_idx] if len(row) > assunto_idx and row[assunto_idx] else None
                observacoes = row[observacoes_idx] if len(row) > observacoes_idx and row[observacoes_idx] else None
                
                # Verificar se já existe um processo com o mesmo número
                if numero_processo and Processo.objects.filter(numero_processo=numero_processo).exists():
                    processos_ignorados += 1
                    erros_detalhados.append(f"Linha {row_num}: Processo com número '{numero_processo}' já existe no sistema.")
                    continue
                
                # Criar o processo com os valores extraídos na ordem correta
                try:
                    setor_value = None
                    if setor:
                        setor_upper = str(setor).upper()
                        if setor_upper == 'CIC' or setor_upper == 'COORDENAÇÃO DE INICIAÇÃO CIENTÍFICA':
                            setor_value = 'CIC'
                        elif setor_upper == 'DPQ' or setor_upper == 'DEPARTAMENTO DE PESQUISA':
                            setor_value = 'DPQ'
                    
                    # Ajuste para Bolsa
                    bolsa_value = None
                    if bolsa:
                        bolsa_upper = str(bolsa).upper()
                        if bolsa_upper == 'SIM' or bolsa_upper == 'S' or bolsa_upper == 'TRUE' or bolsa_upper == '1':
                            bolsa_value = 'Sim'
                        elif bolsa_upper == 'NÃO' or bolsa_upper == 'NAO' or bolsa_upper == 'N' or bolsa_upper == 'FALSE' or bolsa_upper == '0':
                            bolsa_value = 'Não'
                    
                    # Ajuste para Status
                    status_value = None
                    if status:
                        status_upper = str(status).upper()
                        if 'CONCLU' in status_upper or status_upper == 'FINALIZADO':
                            status_value = 'concluido'
                        elif 'ANDAMENTO' in status_upper or 'EM PROCESSO' in status_upper or 'ABERTO' in status_upper:
                            status_value = 'em_andamento'
                    
                    Processo.objects.create(
                        tabela=tabela,
                        nome=nome,
                        matricula=matricula,
                        numero_processo=numero_processo,
                        data_abertura=data_abertura,
                        data_retorno=data_retorno,
                        setor=setor_value,
                        bolsa=bolsa_value,
                        status=status_value,
                        assunto=assunto,
                        observacoes=observacoes
                    )
                    processos_importados += 1
                except Exception as inner_e:
                    # Log detalhado do erro específico desta linha
                    processos_ignorados += 1
                    erros_detalhados.append(f"Linha {row_num}: Erro ao criar processo - {str(inner_e)}")
                    continue
            
            if processos_importados > 0:
                mensagem = f"Processos importados com sucesso! ({processos_importados} importados"
                if processos_ignorados > 0:
                    mensagem += f", {processos_ignorados} ignorados por duplicação ou erro)"
                else:
                    mensagem += ")"
                messages.success(request, mensagem)
            else:
                if processos_ignorados > 0:
                    erro_msg = f"Nenhum processo importado. {processos_ignorados} processos foram ignorados por duplicação ou erro."
                    if erros_detalhados:
                        erro_msg += " Detalhes dos erros foram registrados para o administrador do sistema."
                        # Limitamos a 2 erros para não sobrecarregar a mensagem
                        for i, erro in enumerate(erros_detalhados[:2]):
                            messages.error(request, f"Erro {i+1}: {erro}")
                        if len(erros_detalhados) > 5:
                            messages.error(request, f"... e mais {len(erros_detalhados) - 5} erros.")
                    messages.warning(request, erro_msg)
                else:
                    messages.warning(request, "Nenhum processo válido encontrado no arquivo.")
        except Exception as e:
            messages.error(request, f"Erro ao importar o arquivo: {e}")

    return redirect('tabela_processos', tabela_id=tabela_id)


def login_view(request):
    print(f"DEBUG - Login view - Usuário autenticado: {request.user.is_authenticated}")
    print(f"DEBUG - Login view - Usuário: {request.user}")
    
    if request.user.is_authenticated:
        return redirect('visualizarTabelas')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        print(f"DEBUG - Tentativa de login: {username}")
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            next_page = request.GET.get('next', 'visualizarTabelas')
            messages.success(request, f"Bem-vindo(a), {user.username}!")
            return redirect(next_page)
        else:
            messages.error(request, "Usuário ou senha incorretos. Por favor, tente novamente.")
    
    return render(request, 'login.html')


def logout_view(request):
    logout(request)
    return redirect('login')


@login_required(login_url='login')
@user_passes_test(lambda u: u.is_superuser, login_url='visualizarTabelas')
def alterar_senha_propeg(request):
    try:
        propeg_user = User.objects.get(username='propeg')
    except User.DoesNotExist:
        messages.error(request, "Usuário PROPEG não encontrado no sistema.")
        return redirect('usuarios')

    if request.method == 'POST':
        form = AlterarSenhaPropegForm(request.POST)
        if form.is_valid():
            nova_senha = form.cleaned_data['nova_senha']
            propeg_user.set_password(nova_senha)
            propeg_user.save()
            
            messages.success(request, "Senha do usuário PROPEG alterada com sucesso!")
            return redirect('usuarios')
    else:
        form = AlterarSenhaPropegForm()
    
    return render(request, 'alterar_senha_propeg.html', {'form': form})


@login_required(login_url='login')
def visualizar_auditoria(request):
    acao_filtro = request.GET.get('acao')
    usuario_filtro = request.GET.get('usuario')
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    
    # Query base
    auditorias = Auditoria.objects.all().select_related('usuario', 'processo', 'tabela')
    
    # Aplicar filtros se fornecidos
    if acao_filtro:
        auditorias = auditorias.filter(acao=acao_filtro)
    
    if usuario_filtro:
        auditorias = auditorias.filter(usuario__username__icontains=usuario_filtro)
    
    if data_inicio:
        auditorias = auditorias.filter(data_evento__date__gte=data_inicio)
    
    if data_fim:
        auditorias = auditorias.filter(data_evento__date__lte=data_fim)
    
    # Ordenar por data mais recente
    auditorias = auditorias.order_by('-data_evento')
    
    # Paginação (limitando a 50 registros por página)
    from django.core.paginator import Paginator
    paginator = Paginator(auditorias, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Opções para os filtros
    acoes_disponiveis = Auditoria.AcoesAuditoria.choices
    usuarios_disponiveis = User.objects.filter(acoes_auditadas__isnull=False).distinct()
    
    context = {
        'page_obj': page_obj,
        'acoes_disponiveis': acoes_disponiveis,
        'usuarios_disponiveis': usuarios_disponiveis,
        'acao_filtro': acao_filtro,
        'usuario_filtro': usuario_filtro,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
    }
    
    return render(request, 'auditoria.html', context)
